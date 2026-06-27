import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import numpy as np

class Fluid:
    pass

class HeatExchanger:
    pass

def create_setup_excel(filename="default_setup.xlsx"):
    """
    Create an editable Excel setup template with separate sheets for:
        - f1 (Primary Fluid)
        - f2 (Secondary Fluid)
        - HEX (Heat Exchanger)

    Features:
        - Dropdown menus
        - Default values
        - Descriptions
        - Clean formatting
    """

    # =====================================================
    # Dropdown options
    # =====================================================

    fluids = [
        "air",
        "parahydrogen",
        "water",
        "nitrogen",
        "Hydrotreated mineral oil",
        "custom"
    ]

    directions = [
        "Lx",
        "-Lx",
        "Ly",
        "-Ly",
        "Lz",
        "-Lz"
    ]

    materials = [
        "Aluminum 2219",
        "Aluminum 6061",
        "Copper (C110)",
        "Stainless Steel 304",
        "Inconel 718",
        "Custom"
    ]

    correlations = [
        "tubes (blended Gnielinski)",
        "general (Kays and London correlations)",
        "ASME26 correlations"
    ]

    # =====================================================
    # Workbook
    # =====================================================

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # =====================================================
    # Helper function
    # =====================================================

    def format_sheet(ws):

        headers = [
            "Parameter",
            "Value",
            "Description"
        ]

        for col_idx, header in enumerate(headers, start=1):

            cell = ws.cell(row=1, column=col_idx, value=header)

            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.fill = PatternFill(
                start_color="1F4E78",
                end_color="1F4E78",
                fill_type="solid"
            )

        widths = {
            1: 30,
            2: 35,
            3: 60
        }

        for col_idx, width in widths.items():

            ws.column_dimensions[
                get_column_letter(col_idx)
            ].width = width

        ws.freeze_panes = "A2"
        ws.sheet_view.zoomScale = 120

    # =====================================================
    # Create sheets
    # =====================================================

    ws_f1 = wb.create_sheet("f1")
    ws_f2 = wb.create_sheet("f2")
    ws_hex = wb.create_sheet("HEX")

    ws_lists = wb.create_sheet("DropdownData")

    # =====================================================
    # Dropdown helper data
    # =====================================================

    dropdown_data = {
        "fluids": fluids,
        "directions": directions,
        "materials": materials,
        "correlations": correlations
    }

    range_map = {}

    col = 1

    for name, values in dropdown_data.items():

        col_letter = get_column_letter(col)

        for row_idx, value in enumerate(values, start=1):

            ws_lists.cell(
                row=row_idx,
                column=col,
                value=value
            )

        range_map[name] = (
            f"'DropdownData'!${col_letter}$1:${col_letter}${len(values)}"
        )

        col += 1

    ws_lists.sheet_state = "hidden"

    # =====================================================
    # Sheet formatting
    # =====================================================

    format_sheet(ws_f1)
    format_sheet(ws_f2)
    format_sheet(ws_hex)

    # =====================================================
    # f1 DATA
    # =====================================================

    f1_rows = [
        ["T0", 288.0, "Primary fluid inlet temperature [K]"],
        ["p0", 100000.0, "Primary fluid inlet pressure [Pa]"],
        ["mdot", 12.0, "Primary fluid mass flow rate [kg/s]"],
        ["fluid", "air", "Primary fluid type"]
    ]

    for row_idx, row in enumerate(f1_rows, start=2):

        for col_idx, value in enumerate(row, start=1):

            ws_f1.cell(
                row=row_idx,
                column=col_idx,
                value=value
            )

    # =====================================================
    # f2 DATA
    # =====================================================

    f2_rows = [
        ["T0", 444.0, "Secondary fluid inlet temperature [K]"],
        ["p0", 100000.0, "Secondary fluid inlet pressure [Pa]"],
        ["mdot", 12.0, "Secondary fluid mass flow rate [kg/s]"],
        ["fluid", "water", "Secondary fluid type"]
    ]

    for row_idx, row in enumerate(f2_rows, start=2):

        for col_idx, value in enumerate(row, start=1):

            ws_f2.cell(
                row=row_idx,
                column=col_idx,
                value=value
            )

    # =====================================================
    # HEX DATA
    # =====================================================

    hex_rows = [

        # Geometry
        ["Lx", 0.1, "Heat exchanger X dimension [m]"],
        ["Ly", 0.1, "Heat exchanger Y dimension [m]"],
        ["Lz", 0.1, "Heat exchanger Z dimension [m]"],

        ["rho", 2840, "Heat exchanger material density [kg/m3]"],
        ["k", 120, "Heat exchanger material thermal conductivity [W/m]"],

        ["tw", 0.0005, "Wall thickness [m]"],
        ["tf", 0.0005, "Fin thickness [m]"],

        # Flow
        ["f1_flowdir", "Lx",
         "Primary fluid flow direction"],

        ["f1_n_passes", 1,
         "Primary fluid number of passes"],

        ["f2_flowdir", "Ly",
         "Secondary fluid flow direction"],

        ["f2_n_passes", 1,
         "Secondary fluid number of passes"],

        # Correlations
        ["f1_correlation",
         "ASME26 correlations",
         "Primary fluid correlation"],

        ["f1_l_D_h", "",
         "Primary fluid undisturbed flow length ratio"],

        ["f2_correlation",
         "ASME26 correlations",
         "Secondary fluid correlation"],

        ["f2_l_D_h", "",
         "Secondary fluid undisturbed flow length ratio"],

        # Performance
        ["eta_f", 0.8,
         "Fin thermal efficiency"],

        ["sigma_r", 1.0,
         "Void fraction ratio"],

        ["alpha_r", 1.0,
         "Surface area density ratio"],

        ["chi", 0.2,
         "Compactness / solidity"],

        ["eps", 0.6,
         "Target effectiveness"],

        ["R_h", 1.0,
         "Pressure drop ratio ΔP1/ΔP2"],

        ["R_t", 1.0,
         "Thermal resistance ratio"],
    ]

    for row_idx, row in enumerate(hex_rows, start=2):

        for col_idx, value in enumerate(row, start=1):

            ws_hex.cell(
                row=row_idx,
                column=col_idx,
                value=value
            )

    # =====================================================
    # Create dropdown validations
    # =====================================================

    fluid_validation = DataValidation(
        type="list",
        formula1=range_map["fluids"],
        allow_blank=True
    )

    direction_validation = DataValidation(
        type="list",
        formula1=range_map["directions"],
        allow_blank=True
    )

    material_validation = DataValidation(
        type="list",
        formula1=range_map["materials"],
        allow_blank=True
    )

    correlation_validation = DataValidation(
        type="list",
        formula1=range_map["correlations"],
        allow_blank=True
    )

    # Show dropdown arrows
    fluid_validation.showDropDown = False
    direction_validation.showDropDown = False
    material_validation.showDropDown = False
    correlation_validation.showDropDown = False

    # =====================================================
    # Add validations to sheets
    # =====================================================

    ws_f1.add_data_validation(fluid_validation)
    ws_f2.add_data_validation(fluid_validation)

    ws_hex.add_data_validation(direction_validation)
    ws_hex.add_data_validation(material_validation)
    ws_hex.add_data_validation(correlation_validation)

    # =====================================================
    # Attach dropdowns
    # =====================================================

    # f1
    fluid_validation.add("B5")

    # f2
    fluid_validation.add("B5")

    # HEX
    #material_validation.add("B5")

    direction_validation.add("B9")
    direction_validation.add("B11")

    correlation_validation.add("B13")
    correlation_validation.add("B15")

    # =====================================================
    # Save
    # =====================================================

    wb.save(filename)

    print(f"✅ Default setup template created: {filename}")

def _convert_value(value):

    if value is None:
        return None

    # Numeric values
    if isinstance(value, (int, float, np.integer, np.floating)):
        return np.array([float(value)])

    # Numeric strings
    if isinstance(value, str):

        stripped = value.strip()

        try:
            return np.array([float(stripped)])

        except:
            pass

    return value

def _sheet_to_object(ws, obj_class=Fluid):
    """
    Convert Excel sheet into object.

    Assumes:
        Column A -> Property
        Column B -> Value
    """

    obj = obj_class()

    for row in ws.iter_rows(min_row=2, values_only=True):

        param = row[0]
        value = row[1]

        if param is None:
            continue

        converted = _convert_value(value)

        setattr(obj, param, converted)

    return obj

def load_setup_from_excel(filename):
    """
    Load setup from Excel workbook.

    Required sheets:
        - f1
        - f2
        - HEX

    Returns
    -------
    filename : str
    setup : dict
    """

    wb = load_workbook(filename)

    required_sheets = [
        "f1",
        "f2",
        "HEX"
    ]

    for s in required_sheets:

        if s not in wb.sheetnames:
            raise ValueError(f"Missing required sheet: {s}")

    # -------------------------------------------------
    # Load objects
    # -------------------------------------------------

    f1 = _sheet_to_object(
        wb["f1"],
        obj_class=Fluid
    )

    f2 = _sheet_to_object(
        wb["f2"],
        obj_class=Fluid
    )

    HEX = _sheet_to_object(
        wb["HEX"],
        obj_class=HeatExchanger
    )

    # -------------------------------------------------
    # Build setup dict
    # -------------------------------------------------

    setup = {
        "f1": f1,
        "f2": f2,
        "HEX": HEX
    }

    print(f"📂 Loaded Excel setup: {filename}")

    print(
        f"   f1 fluid: {getattr(f1, 'fluid', 'unknown')}"
    )

    print(
        f"   f2 fluid: {getattr(f2, 'fluid', 'unknown')}"
    )

    return setup

def save_output_objects_to_excel(filename, setup):
    """
    Save output objects into NEW sheets in an existing Excel file.

    Creates/overwrites:
        - f1_output
        - f2_output
        - HEX_output

    Each sheet contains:
        Column A -> Property
        Column B -> Value
    """
    f1 = setup["f1"]
    f2 = setup["f2"]
    HEX = setup["HEX"]

    # =====================================================
    # Load existing workbook
    # =====================================================

    wb = load_workbook(filename)

    # =====================================================
    # Remove old output sheets if they exist
    # =====================================================

    output_sheets = [
        "f1_output",
        "f2_output",
        "HEX_output"
    ]

    for s in output_sheets:

        if s in wb.sheetnames:
            del wb[s]

    # =====================================================
    # Helper function
    # =====================================================

    def clean_value(v):

        # numpy arrays
        if isinstance(v, np.ndarray):

            if v.size == 1:
                return v.item()

            return str(v.tolist())

        # numpy scalars
        elif isinstance(v, (
            np.float64,
            np.float32,
            np.int64,
            np.int32
        )):

            return v.item()

        # lists/tuples
        elif isinstance(v, (list, tuple)):

            return str(v)

        return v

    # =====================================================
    # Helper to create output sheet
    # =====================================================

    def write_object_sheet(sheet_name, obj):

        ws = wb.create_sheet(sheet_name)

        # -------------------------------------------------
        # Header
        # -------------------------------------------------

        headers = ["Property", "Value"]

        for col_idx, header in enumerate(headers, start=1):

            cell = ws.cell(
                row=1,
                column=col_idx,
                value=header
            )

            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.fill = PatternFill(
                start_color="1F4E78",
                end_color="1F4E78",
                fill_type="solid"
            )

        # -------------------------------------------------
        # Write object properties
        # -------------------------------------------------

        for row_idx, (key, value) in enumerate(
            obj.__dict__.items(),
            start=2
        ):

            ws.cell(
                row=row_idx,
                column=1,
                value=key
            )

            ws.cell(
                row=row_idx,
                column=2,
                value=clean_value(value)
            )

        # -------------------------------------------------
        # Formatting
        # -------------------------------------------------

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 50

        ws.freeze_panes = "A2"

        ws.sheet_view.zoomScale = 120

    # =====================================================
    # Create output sheets
    # =====================================================

    write_object_sheet("f1_output", f1)
    write_object_sheet("f2_output", f2)
    write_object_sheet("HEX_output", HEX)

    # =====================================================
    # Save workbook
    # =====================================================

    wb.save(filename)

    print(f"✅ Output sheets saved to: {filename}")