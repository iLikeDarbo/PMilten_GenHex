import numpy as np
import pickle
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# -------------------------------
# Define data classes
# -------------------------------
class Fluid:
    """Fluid parameters."""
    pass

class HeatExchanger:
    """Geometric, material, and general setup."""
    pass

# -------------------------------
# Helper functions
# -------------------------------
def get_input(prompt, cast_type=float, default=None):
    """Prompt user safely with type conversion and default values."""
    while True:
        user_input = input(f"{prompt} [{default if default is not None else ''}]: ").strip()
        if not user_input:
            return default
        try:
            return cast_type(user_input)
        except ValueError:
            print(f"⚠️ Please enter a valid {cast_type.__name__} value.")

def choose_direction(prompt, default="Lx"):
    """Prompt user to select a flow direction from Lx, Ly, or Lz."""
    valid = ["Lx", "-Lx", "Ly", "-Ly", "Lz", "-Lz"]
    while True:
        user_input = input(f"{prompt} ({'/'.join(valid)}) [{default}]: ").strip()
        if not user_input:
            return default
        if user_input in valid:
            return user_input
        print("⚠️ Invalid choice. Please type Lx, -Lx, Ly, -Lx, Lz or -Lz.")

def choose_material(prev=None):
    """Prompt user to select a material from a list."""
    materials = {
        "1": ("Aluminum 2219", 120, 2840),
        "2": ("Aluminum 6061", 167, 2700),
        "3": ("Copper (C110)", 390, 8960),
        "4": ("Stainless Steel 304", 16, 8000),
        "5": ("Inconel 718", 11, 8190),
        "6": ("Custom", None, None),
    }

    print("\n🔩 Available Materials:")
    for key, (name, k, rho) in materials.items():
        if k and rho:
            print(f"  {key}. {name:20s} — k={k} W/m·K, ρ={rho} kg/m³")
        else:
            print(f"  {key}. {name}")

    default_choice = prev.get("material_choice", "1") if prev else "1"
    choice = input(f"Select material [1–6] [{default_choice}]: ").strip() or default_choice

    if choice not in materials:
        print("⚠️ Invalid selection, using 'Custom'.")
        choice = "6"

    name, k, rho = materials[choice]

    if name == "Custom":
        k = get_input("Enter thermal conductivity [W/m·K]", float, prev.get("k", 120) if prev else 120)
        rho = get_input("Enter density [kg/m³]", float, prev.get("rho", 1250) if prev else 1250)

    return name, k, rho, choice

def choose_fluid(prompt, prev=None, default="air"):
    """Prompt user to select a fluid or enter a custom name."""
    fluids = ["air", "parahydrogen", "water", "nitrogen", "Hydrotreated mineral oil", "custom"]
    prev_fluid = prev.get(prompt.lower().replace(" ", "_"), default) if prev else default

    while True:
        print("\n💧 Available fluids:")
        for f in fluids:
            print("  -", f)
        choice = input(f"{prompt} [{prev_fluid}]: ").strip().lower()
        if not choice:
            choice = prev_fluid
        if choice in fluids:
            break
        print("⚠️ Invalid choice, please select from air, parahydrogen, water, or custom.")

    if choice == "custom":
        choice = input("Enter custom fluid name: ").strip()

    return choice

def choose_correlation(fluid_name, prev=None, prev_l_D_h=None, default="tubes (blended Gnielinski)"):
    """Prompt user to select a correlation for aerothermal performance."""
    options = ["tubes (blended Gnielinski)", "general (Kays and London correlations)", "ASME26 correlations"]
    # prev_choice = prev.get("correlation", default) if prev else default
    prev_choice = prev if prev else default

    print(f"\n📊 Select correlation for {fluid_name}:")
    for i, opt in enumerate(options, start=1):
        print(f"  {i}. {opt}")

    while True:
        choice = input(f"Enter choice [1-{len(options)}] [{prev_choice}]: ").strip()
        if not choice:
            selected = prev_choice
            break
        elif choice.isdigit() and 1 <= int(choice) <= len(options):
            selected = options[int(choice) - 1]
            break
        else:
            print(f"⚠️ Invalid choice. Pick 1 or 2.")

    # If general Kays & London is selected, ask for undisturbed flow length
    if selected == "general (Kays and London correlations)" or selected == "ASME26 correlations" :
        f_l = get_input(f"Enter ratio of undisturbed flow length to hydraulic diameter for {fluid_name} []", float,
                        prev_l_D_h if prev_l_D_h else None)

    return selected, f_l

def load_previous_setup(file_path="setup.pkl"):
    """Load an existing setup if available."""
    if os.path.exists(file_path):
        with open(file_path, "rb") as f:
            prev_setup = pickle.load(f)
        print("📂 Previous setup detected — using saved values as defaults.\n")
        return prev_setup
    else:
        print("🆕 No previous setup found — using standard defaults.\n")
        return None

def choose_filename(default="setup.pkl", folder="."):
    """Select or create a setup file, with option to overwrite existing ones.
       Returns (filename, prev) where prev is the loaded setup or None."""
    # List all existing .pkl setup files
    setups = [f for f in os.listdir(folder) if f.endswith(".pkl")]
    print("\n📁 Existing setup files:")
    if setups:
        for i, s in enumerate(setups, start=1):
            print(f"  {i}. {s}")
    else:
        print("  (none found)")

    print("\nOptions:")
    print("  [Enter] → create a new setup")
    if setups:
        print("  [number] → overwrite or reuse an existing setup")

    # Ask for choice
    choice = input("\nSelect setup to overwrite or press Enter for new: ").strip()

    prev = None
    if choice == "":
        # User wants to create a new setup
        filename = input(f"Enter new filename [{default}]: ").strip()
        filename = filename if filename else default
    elif choice.isdigit() and 1 <= int(choice) <= len(setups):
        # User selected an existing file
        filename = setups[int(choice) - 1]
        confirm = input(f"⚠️ Overwrite or reuse {filename}? [y/N]: ").strip().lower()
        if confirm != "y":
            print("Operation cancelled. No setup saved.")
            return None, None
        # Try loading previous setup
        try:
            with open(os.path.join(folder, filename), "rb") as f:
                prev = pickle.load(f)
            print(f"✅ Loaded previous setup: {filename}")
        except Exception as e:
            print(f"⚠️ Could not load existing setup ({e}). Starting fresh.")
            prev = None
    else:
        print("Invalid choice. Using default.")
        filename = default

    # Ensure .pkl extension
    if not filename.endswith(".pkl"):
        filename += ".pkl"

    filepath = os.path.join(folder, filename)
    print(f"💾 Selected setup file: {filename}")
    return filepath, prev

def object_to_dict(obj):
    """
    Convert an object into a clean dictionary suitable for Excel export.
    Handles numpy arrays and scalar conversion.
    """
    out = {}

    for key, value in obj.__dict__.items():

        # Convert numpy arrays
        if isinstance(value, np.ndarray):
            if value.size == 1:
                out[key] = value.item()
            else:
                out[key] = ", ".join(map(str, value.tolist()))

        # Convert numpy scalars
        elif isinstance(value, (np.floating, np.integer)):
            out[key] = value.item()

        else:
            out[key] = value

    return out

def save_setup_to_excel(f1, f2, HEX, filename="setup.xlsx"):
    """
    Save setup objects (f1, f2, HEX) into an editable Excel file.

    Structure:
        Sheet 1 -> Primary Fluid (f1)
        Sheet 2 -> Secondary Fluid (f2)
        Sheet 3 -> Heat Exchanger (HEX)

    Parameters
    ----------
    f1 : object
        Primary fluid object
    f2 : object
        Secondary fluid object
    HEX : object
        Heat exchanger object
    filename : str
        Output Excel filename
    """

    # Convert objects to dictionaries
    f1_dict = object_to_dict(f1)
    f2_dict = object_to_dict(f2)
    hex_dict = object_to_dict(HEX)

    # Create DataFrames
    df_f1 = pd.DataFrame(list(f1_dict.items()), columns=["Parameter", "Value"])
    df_f2 = pd.DataFrame(list(f2_dict.items()), columns=["Parameter", "Value"])
    df_hex = pd.DataFrame(list(hex_dict.items()), columns=["Parameter", "Value"])

    # Write Excel file
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df_f1.to_excel(writer, sheet_name="PrimaryFluid", index=False)
        df_f2.to_excel(writer, sheet_name="SecondaryFluid", index=False)
        df_hex.to_excel(writer, sheet_name="HeatExchanger", index=False)

    # Formatting
    wb = load_workbook(filename)

    for ws in wb.worksheets:

        # Bold headers
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Auto column widths
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)

            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = max_length + 4
            ws.column_dimensions[column].width = adjusted_width

    wb.save(filename)

    print(f"✅ Excel setup file saved as: {filename}")

# -------------------------------
# Main interactive function
# -------------------------------
def UI_main():
    print("🔧 HEAT EXCHANGER SETUP CONFIGURATION TOOL")
    print("------------------------------------------\n")

    filename, prev = choose_filename()

    #prev = load_previous_setup()
    f1 = Fluid()
    f2 = Fluid()
    HEX = HeatExchanger()

    # -------------------------------
    # Primary fluid setup
    # -------------------------------
    f1.T0 = np.array([get_input("Primary fluid temperature [K]", float, getattr(prev["f1"], "T0", [288])[0] if prev else 288)])
    f1.p0 = np.array([get_input("Primary fluid pressure [Pa]", float, getattr(prev["f1"], "p0", [100000])[0] if prev else 100000)])
    f1.mdot = np.array([get_input("Primary fluid mass flow [kg/s]", float, getattr(prev["f1"], "mdot", [32])[0] if prev else 32)])
    f1.fluid = choose_fluid("Select primary fluid", prev["f1"].__dict__ if prev else None, getattr(prev["f1"], "fluid", "air") if prev else "air")

    # -------------------------------
    # Secondary fluid setup
    # -------------------------------
    f2.T0 = np.array([get_input("Secondary fluid temperature [K]", float, getattr(prev["f2"], "T0", [444])[0] if prev else 444)])
    f2.p0 = np.array([get_input("Secondary fluid pressure [Pa]", float, getattr(prev["f2"], "p0", [100000])[0] if prev else 100000)])
    f2.mdot = np.array([get_input("Secondary fluid mass flow [kg/s]", float, getattr(prev["f2"], "mdot", [32])[0] if prev else 32)])
    f2.fluid = choose_fluid("Select secondary fluid", prev["f2"].__dict__ if prev else None, getattr(prev["f2"], "fluid", "water") if prev else "water")

    # -------------------------------
    # HEX geometry & material
    # -------------------------------
    HEX.Lx = get_input("Length X [m]", float, getattr(prev["HEX"], "Lx", 0.1) if prev else 0.1)
    HEX.Ly = get_input("Length Y [m]", float, getattr(prev["HEX"], "Ly", 0.1) if prev else 0.1)
    HEX.Lz = get_input("Length Z [m]", float, getattr(prev["HEX"], "Lz", 0.1) if prev else 0.1)

    mat_name, k, rho, choice = choose_material(prev["HEX"].__dict__ if prev else None)
    HEX.material = mat_name
    HEX.k = k
    HEX.rho = rho
    HEX.material_choice = choice

    HEX.tw = get_input("Wall thickness [m]", float, getattr(prev["HEX"], "tw", 0.0005) if prev else 0.0005)
    HEX.tf = get_input("Fin thickness [m]", float, getattr(prev["HEX"], "tf", 0.0005) if prev else 0.0005)

    # Flow direction and passes
    HEX.f1_flowdir = choose_direction("Primary fluid flow direction", getattr(prev["HEX"], "f1_flowdir", "Lx") if prev else "Lx")
    HEX.f1_n_passes = get_input("Primary fluid number of passes", int, getattr(prev["HEX"], "f1_n_passes", 1) if prev else 1)

    HEX.f2_flowdir = choose_direction("Secondary fluid flow direction", getattr(prev["HEX"], "f2_flowdir", "Ly") if prev else "Ly")
    HEX.f2_n_passes = get_input("Secondary fluid number of passes", int, getattr(prev["HEX"], "f2_n_passes", 1) if prev else 1)

    # Aerothermal correlations
    # For primary fluid
    # HEX.f1_correlation, HEX.f1_l = choose_correlation("Primary fluid", prev["f1"].__dict__ if prev else None)
    HEX.f1_correlation, HEX.f1_l_D_h = choose_correlation("Primary fluid", getattr(prev["HEX"], "f1_correlation") if prev else None, getattr(prev["HEX"], "f1_l_D_h") if prev else None)

    # For secondary fluid
    HEX.f2_correlation, HEX.f2_l_D_h = choose_correlation("Secondary fluid", getattr(prev["HEX"], "f2_correlation") if prev else None, getattr(prev["HEX"], "f2_l_D_h") if prev else None)

    # Estimate fin efficiency
    HEX.f1_eta_f = np.array([get_input("Fin thermal efficiency ($\eta_f$)", float, getattr(prev["HEX"], "f1_eta_f", 1.0)[0] if prev else 1.0)])
    HEX.f2_eta_f = HEX.f1_eta_f

    # Geometrical parameters & objective
    HEX.sigma_r = np.array([get_input("Void fraction ratio (σ_r)", float, getattr(prev["HEX"], "sigma_r", 1.0)[0] if prev else 1.0)])
    HEX.alpha_r = np.array([get_input("Surface area density ratio (α_r)", float, getattr(prev["HEX"], "alpha_r", 1.0)[0] if prev else 1.0)])
    HEX.chi = np.array([get_input("Solidity / compactness (χ)", float, getattr(prev["HEX"], "chi", 0.2)[0] if prev else 0.2)])
    HEX.eps = np.array([get_input("Target effectiveness (ε)", float, getattr(prev["HEX"], "eps", 0.8)[0] if prev else 1.0)])
    HEX.dP_ratio = np.array([get_input("Pressure drop ratio (ΔP1/ΔP2)", float, getattr(prev["HEX"], "dP_ratio", 1.0)[0] if prev else 1.0)])

    # -------------------------------
    # Save all objects in one pickle
    # -------------------------------
    setup = {"f1": f1, "f2": f2, "HEX": HEX}
    with open(filename, "wb") as f:
        pickle.dump(setup, f)

    print("\n✅ Setup created and saved to"+ filename+":")
    print("  → f1, f2, and HEX objects are included.\n")

    excel_filename = filename.replace(".pkl", ".xlsx")
    save_setup_to_excel(f1, f2, HEX, excel_filename)

# -------------------------------
# Entry point
# -------------------------------
if __name__ == "__main__":
    UI_main()
